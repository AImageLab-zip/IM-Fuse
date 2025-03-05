import pandas as pd
import json
import os
from openpyxl import load_workbook


def aggregate_specific_summaries(base_path):
    all_data = {}
    
    # 遍历base_path下的所有文件夹
    for folder_name in os.listdir(base_path):
        folder_path = os.path.join(base_path, folder_name)
        if os.path.isdir(folder_path):
            summary_file = os.path.join(folder_path, 'summary.json')
            if os.path.exists(summary_file):
                with open(summary_file, 'r') as file:
                    data = json.load(file)
                    # 提取特定的部分
                    extracted_data = {
                        'foreground_mean': data.get('foreground_mean', {}),
                        'mean': data.get('mean', {})
                    }
                    all_data[folder_name] = extracted_data
    
    # 将汇总后的数据写入新的JSON文件
    with open(os.path.join(base_path, 'aggregated_specific_summary.json'), 'w') as file:
        json.dump(all_data, file, indent=4)



def load_data_from_json(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)
    rows = []
    for folder_name, contents in data.items():
        mask_index = folder_name.split('_')[-1]
        row = {'Validation Set': mask_index}
        #row['Foreground Mean Dice'] = contents['foreground_mean']['Dice'] if 'foreground_mean' in contents else None
        for key, value in contents.get('mean', {}).items():
            #row[f'{key} Hausdorff_distance'] = value['Hausdorff_distance']
            row[f'{key} Dice'] = value['Dice']
        rows.append(row)
    return pd.DataFrame(rows)

def append_or_create_sheet(json_file_path, excel_file_path):
    # 读取或创建Excel写入器
    if os.path.exists(excel_file_path):
        book = load_workbook(excel_file_path)
        writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
    else:
        writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')

    # 加载数据
    df = load_data_from_json(json_file_path)
    sheet_name = os.path.basename(json_file_path).replace('.json', '')

    # 写入数据到Excel文件
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    writer.close() # 释放资源


# 调用函数
base_path = '/data_hdd/users/zengzhilin/nnUNetV2_trained_models/Dataset520_BraTS2020/nnUNetTrainerMissingReconBaseline__nnUNetPlans__3d_fullres_multiencoder_recon_base/fold_0'
aggregate_specific_summaries(base_path)
# 示例调用
append_or_create_sheet(os.path.join(base_path, 'aggregated_specific_summary.json'), 
                       '/data_hdd/users/zengzhilin/nnUNet/Aggregated_Results_recon__base_Dice_2020.xlsx')

